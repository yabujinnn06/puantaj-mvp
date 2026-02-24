from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from numbers import Number
from typing import Literal

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import quote_sheetname
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import AttendanceEvent, AttendanceType, Department, Employee, WorkRule
from app.schemas import MonthlyEmployeeResponse
from app.services.attendance import _attendance_timezone
from app.services.monthly import calculate_employee_monthly
from app.services.monthly_calc import calculate_work_and_overtime

ExportMode = Literal["employee", "department", "all", "date_range"]
RangeSheetMode = Literal["consolidated", "employee_sheets", "department_sheets"]

DEFAULT_DAILY_MINUTES_PLANNED = 540
DEFAULT_BREAK_MINUTES = 60
DURATION_NUMBER_FORMAT = "[h]:mm"

DAILY_HEADERS = [
    "Tarih",
    "Giriş",
    "Çıkış",
    "Saat Aralığı",
    "Çalışma Süresi",
    "Mola",
    "Net Süre",
    "Net Süre (dk)",
    "Plan Üstü Süre",
    "Plan Üstü Süre (dk)",
    "Fazla Sürelerle Çalışma",
    "Yasal Fazla Mesai",
    "Yasal Fazla Mesai (dk)",
    "Gün Tipi",
    "Çalışıldı mı",
    "Durum",
    "Bayraklar",
]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="0B4F73")
SUBHEADER_FILL = PatternFill(fill_type="solid", fgColor="DCEBF3")
META_LABEL_FILL = PatternFill(fill_type="solid", fgColor="EAF4F9")
META_VALUE_FILL = PatternFill(fill_type="solid", fgColor="F8FCFF")
ZEBRA_FILL = PatternFill(fill_type="solid", fgColor="F7FBFE")
WARNING_FILL = PatternFill(fill_type="solid", fgColor="FFF3CD")
ALERT_FILL = PatternFill(fill_type="solid", fgColor="FDE2E4")
SUCCESS_FILL = PatternFill(fill_type="solid", fgColor="E6F4EA")
SUMMARY_FILL = PatternFill(fill_type="solid", fgColor="E9F1F7")
README_PANEL_FILL = PatternFill(fill_type="solid", fgColor="F3F8FC")
NEUTRAL_PANEL_FILL = PatternFill(fill_type="solid", fgColor="EEF2F7")

HEADER_FONT = Font(bold=True, color="FFFFFF")
BOLD_FONT = Font(bold=True, color="0F172A")
TITLE_FONT = Font(bold=True, color="0B4F73", size=14)
MUTED_FONT = Font(color="334155")

THIN_SIDE = Side(style="thin", color="D5E2EC")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def _to_excel_datetime(value: datetime | None) -> datetime | None:
    if value is None:
        return None
    if value.tzinfo is None:
        return value
    return value.astimezone(timezone.utc).replace(tzinfo=None)


def _date_range_bounds_utc(start_date: date, end_date: date) -> tuple[datetime, datetime]:
    tz = _attendance_timezone()
    start_dt = datetime.combine(start_date, datetime.min.time(), tzinfo=tz).astimezone(timezone.utc)
    end_dt = datetime.combine(end_date + timedelta(days=1), datetime.min.time(), tzinfo=tz).astimezone(timezone.utc)
    return start_dt, end_dt


def _local_day_from_ts(value: datetime) -> date:
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    return value.astimezone(_attendance_timezone()).date()


def _minutes_to_hhmm(minutes: int) -> str:
    value = max(0, int(minutes))
    hours = value // 60
    mins = value % 60
    return f"{hours:02d}:{mins:02d}"


def _minutes_to_excel_duration(minutes: int) -> float:
    return max(0, int(minutes)) / 1440.0


def _is_special_day(*, day_date: date, leave_type: object, flags: list[str]) -> bool:
    if str(leave_type or "").upper() == "PUBLIC_HOLIDAY":
        return True
    if day_date.weekday() == 6 and "PUBLIC_HOLIDAY" in flags:
        return True
    return "PUBLIC_HOLIDAY" in flags


def _day_type_label(*, day_date: date, leave_type: object, flags: list[str]) -> str:
    if _is_special_day(day_date=day_date, leave_type=leave_type, flags=flags):
        return "Özel Gün"
    if day_date.weekday() == 6:
        return "Pazar"
    return "Hafta İçi"


def _was_worked_day(*, worked_minutes: int, check_in: datetime | None, check_out: datetime | None) -> bool:
    if worked_minutes > 0:
        return True
    return check_in is not None or check_out is not None


def _apply_print_layout(ws: Worksheet, *, header_row: int | None = None) -> None:
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.sheet_view.zoomScale = 90
    if header_row is not None and header_row > 0:
        ws.print_title_rows = f"{header_row}:{header_row}"


def _build_summary_row_from_daily_facts(
    *,
    employee_id: int,
    employee_name: str,
    department_name: str,
    rows: list[dict[str, object]],
) -> dict[str, object]:
    worked_minutes = sum(int(item["worked_minutes"]) for item in rows)
    plan_overtime_minutes = sum(int(item.get("plan_overtime_minutes", 0)) for item in rows)
    extra_work_minutes = sum(int(item["extra_work_minutes"]) for item in rows)
    overtime_minutes = sum(int(item["overtime_minutes"]) for item in rows)
    incomplete_days = sum(1 for item in rows if str(item["status"]).upper() == "INCOMPLETE")
    worked_days = sum(1 for item in rows if bool(item["worked_flag"]))
    sunday_worked_days = sum(1 for item in rows if bool(item["worked_flag"]) and item["day_type"] == "Pazar")
    special_worked_days = sum(1 for item in rows if bool(item["worked_flag"]) and item["day_type"] == "Özel Gün")
    sunday_worked_minutes = sum(
        int(item["worked_minutes"])
        for item in rows
        if bool(item["worked_flag"]) and item["day_type"] == "Pazar"
    )
    special_worked_minutes = sum(
        int(item["worked_minutes"])
        for item in rows
        if bool(item["worked_flag"]) and item["day_type"] == "Özel Gün"
    )

    return {
        "employee_id": employee_id,
        "employee_name": employee_name,
        "department_name": department_name,
        "worked_minutes": worked_minutes,
        "plan_overtime_minutes": plan_overtime_minutes,
        "extra_work_minutes": extra_work_minutes,
        "overtime_minutes": overtime_minutes,
        "incomplete_days": incomplete_days,
        "annual_overtime_minutes": overtime_minutes,
        "worked_days": worked_days,
        "sunday_worked_days": sunday_worked_days,
        "special_worked_days": special_worked_days,
        "sunday_worked_minutes": sunday_worked_minutes,
        "special_worked_minutes": special_worked_minutes,
    }


def _group_summary_rows_by_department(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    department_map: dict[str, dict[str, object]] = {}
    member_map: dict[str, set[int]] = defaultdict(set)
    for row in rows:
        department_name = str(row["department_name"] or "-")
        if department_name not in department_map:
            department_map[department_name] = {
                "department_name": department_name,
                "employee_count": 0,
                "worked_minutes": 0,
                "plan_overtime_minutes": 0,
                "extra_work_minutes": 0,
                "overtime_minutes": 0,
                "incomplete_days": 0,
                "worked_days": 0,
                "sunday_worked_days": 0,
                "special_worked_days": 0,
                "sunday_worked_minutes": 0,
                "special_worked_minutes": 0,
            }

        bucket = department_map[department_name]
        employee_id = int(row["employee_id"])
        member_map[department_name].add(employee_id)
        bucket["worked_minutes"] = int(bucket["worked_minutes"]) + int(row["worked_minutes"])
        bucket["plan_overtime_minutes"] = int(bucket["plan_overtime_minutes"]) + int(row.get("plan_overtime_minutes", 0))
        bucket["extra_work_minutes"] = int(bucket["extra_work_minutes"]) + int(row["extra_work_minutes"])
        bucket["overtime_minutes"] = int(bucket["overtime_minutes"]) + int(row["overtime_minutes"])
        bucket["incomplete_days"] = int(bucket["incomplete_days"]) + int(row["incomplete_days"])
        bucket["worked_days"] = int(bucket["worked_days"]) + int(row["worked_days"])
        bucket["sunday_worked_days"] = int(bucket["sunday_worked_days"]) + int(row["sunday_worked_days"])
        bucket["special_worked_days"] = int(bucket["special_worked_days"]) + int(row["special_worked_days"])
        bucket["sunday_worked_minutes"] = int(bucket["sunday_worked_minutes"]) + int(row["sunday_worked_minutes"])
        bucket["special_worked_minutes"] = int(bucket["special_worked_minutes"]) + int(row["special_worked_minutes"])

    ordered_rows = []
    for department_name in sorted(department_map):
        row = department_map[department_name]
        row["employee_count"] = len(member_map[department_name])
        ordered_rows.append(row)
    return ordered_rows

def _build_daily_legal_breakdown(
    report: MonthlyEmployeeResponse,
    *,
    contract_weekly_minutes: int | None,
) -> dict[date, tuple[int, int]]:
    _ = contract_weekly_minutes
    result: dict[date, tuple[int, int]] = {}
    for day in report.days:
        plan_overtime_minutes = max(
            0,
            int(getattr(day, "plan_overtime_minutes", getattr(day, "overtime_minutes", 0))),
        )
        result[day.date] = (0, plan_overtime_minutes)
    return result


def _style_header(ws: Worksheet, row: int = 1) -> None:
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws: Worksheet) -> None:
    for column_cells in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        max_len = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.coordinate in ws.merged_cells:
                continue
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)


def _append_spacer_row(ws: Worksheet) -> None:
    # `ws.append([])` does not always advance `max_row`; write a blank string cell for stable row indexing.
    ws.append([""])


def _merge_title(ws: Worksheet, row: int, text: str, *, end_col: int | None = None) -> None:
    max_col = max(1, int(end_col) if end_col is not None else ws.max_column)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _finalize_visual_scope(ws: Worksheet, *, used_max_col: int) -> None:
    ws.sheet_view.showGridLines = False
    safe_col = max(1, used_max_col)
    safe_row = max(1, ws.max_row)
    ws.print_area = f"A1:{get_column_letter(safe_col)}{safe_row}"


def _style_metadata_rows(ws: Worksheet, *, start_row: int, end_row: int) -> None:
    for row_idx in range(start_row, end_row + 1):
        label_cell = ws.cell(row=row_idx, column=1)
        value_cell = ws.cell(row=row_idx, column=2)
        label_cell.font = BOLD_FONT
        label_cell.fill = META_LABEL_FILL
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        label_cell.border = THIN_BORDER

        value_cell.font = MUTED_FONT
        value_cell.fill = META_VALUE_FILL
        value_cell.alignment = Alignment(horizontal="left", vertical="center")
        value_cell.border = THIN_BORDER


def _is_hhmm_value(value: object) -> bool:
    if not isinstance(value, str):
        return False
    if len(value) != 5 or value[2] != ":":
        return False
    hh, mm = value.split(":")
    return hh.isdigit() and mm.isdigit()


def _style_table_region(
    ws: Worksheet,
    *,
    header_row: int,
    data_start_row: int,
    data_end_row: int,
    status_col_name: str = "Durum",
    flags_col_name: str = "Bayraklar",
) -> None:
    if data_end_row < data_start_row:
        ws.freeze_panes = f"A{header_row + 1}"
        return

    table_max_col = _resolve_table_max_col(ws, header_row=header_row)
    header_map: dict[str, int] = {}
    for col_idx in range(1, table_max_col + 1):
        header_value = ws.cell(row=header_row, column=col_idx).value
        if isinstance(header_value, str) and header_value.strip():
            header_map[header_value] = col_idx

    ws.freeze_panes = f"A{header_row + 1}"

    status_col = header_map.get(status_col_name)
    flags_col = header_map.get(flags_col_name)
    overtime_cols = [
        idx
        for name, idx in header_map.items()
        if (
            "Fazla Mesai" in name
            or "Fazla Surelerle" in name
            or "Fazla S\u00fcrelerle" in name
            or "Plan Ustu" in name
            or "Plan \u00dcst\u00fc" in name
        )
    ]

    for row_idx in range(data_start_row, data_end_row + 1):
        status_value = ws.cell(row=row_idx, column=status_col).value if status_col else None
        flags_value = ws.cell(row=row_idx, column=flags_col).value if flags_col else None

        if isinstance(status_value, str) and status_value.upper() in {"INCOMPLETE", "LEAVE", "OFF"}:
            row_fill = WARNING_FILL
        elif isinstance(status_value, str) and status_value.upper() in {"OK", "FINISHED"}:
            row_fill = PatternFill(fill_type=None)
        else:
            row_fill = PatternFill(fill_type=None)

        if row_fill.fill_type is None and row_idx % 2 == 0:
            row_fill = ZEBRA_FILL

        for col_idx in range(1, table_max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            if row_fill.fill_type:
                cell.fill = row_fill
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif _is_hhmm_value(cell.value):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        if flags_col and flags_value not in {None, "", "-"}:
            flag_cell = ws.cell(row=row_idx, column=flags_col)
            flag_cell.fill = ALERT_FILL
            flag_cell.font = Font(bold=True, color="9F1239")

        for overtime_col in overtime_cols:
            overtime_cell = ws.cell(row=row_idx, column=overtime_col)
            if overtime_cell.value not in {None, "", "00:00", 0}:
                overtime_cell.fill = SUCCESS_FILL
                overtime_cell.font = Font(bold=True, color="166534")

    _apply_conditional_formatting(
        ws,
        header_row=header_row,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        header_map=header_map,
        status_col_name=status_col_name,
        flags_col_name=flags_col_name,
    )
    table_added = _add_interactive_table(
        ws,
        header_row=header_row,
        data_end_row=data_end_row,
        max_col=table_max_col,
    )
    if table_added:
        ws.auto_filter.ref = None
    else:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(table_max_col)}{data_end_row}"


def _apply_duration_formats(
    ws: Worksheet,
    *,
    header_row: int,
    data_start_row: int,
    data_end_row: int,
    header_names: list[str],
) -> None:
    if data_end_row < data_start_row:
        return
    header_index: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col_idx).value
        if isinstance(value, str):
            header_index[value] = col_idx

    target_columns = {header_index[name] for name in header_names if name in header_index}
    for header_name, col_idx in header_index.items():
        if header_name.endswith("(dk)"):
            continue
        if "[h:mm]" in header_name or "Süre" in header_name or "Mesai" in header_name:
            target_columns.add(col_idx)
    for col_idx in sorted(target_columns):
        for row_idx in range(data_start_row, data_end_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, Number):
                cell.number_format = DURATION_NUMBER_FORMAT
            elif isinstance(cell.value, str):
                normalized = cell.value.strip().replace(",", ".")
                try:
                    parsed = float(normalized)
                except ValueError:
                    continue
                cell.value = parsed
                cell.number_format = DURATION_NUMBER_FORMAT


def _safe_sheet_title(title: str, fallback: str) -> str:
    cleaned = "".join(ch for ch in title if ch not in ['\\', '/', '*', '?', ':', '[', ']']).strip()
    if not cleaned:
        cleaned = fallback
    return cleaned[:31]


def _set_internal_sheet_link(cell: object, sheet_title: str) -> None:
    if not hasattr(cell, "hyperlink"):
        return
    target = f"#{quote_sheetname(sheet_title)}!A1"
    setattr(cell, "hyperlink", target)
    setattr(cell, "style", "Hyperlink")


def _table_range_ref(ws: Worksheet, *, header_row: int, data_end_row: int, max_col: int) -> str:
    return f"A{header_row}:{get_column_letter(max_col)}{data_end_row}"


def _existing_table_names(ws: Worksheet) -> set[str]:
    names: set[str] = set()
    workbook = ws.parent
    for sheet in workbook.worksheets:
        names.update(str(name) for name in sheet.tables.keys())
    return names


def _build_unique_table_name(ws: Worksheet, *, seed: str) -> str:
    cleaned = "".join(ch if ch.isalnum() else "_" for ch in seed)
    if not cleaned:
        cleaned = "TABLE"
    if cleaned[0].isdigit():
        cleaned = f"T_{cleaned}"
    base = cleaned[:48]
    existing = _existing_table_names(ws)
    candidate = base
    counter = 1
    while candidate in existing:
        suffix = f"_{counter}"
        candidate = f"{base[: max(1, 48 - len(suffix))]}{suffix}"
        counter += 1
    return candidate


def _resolve_table_max_col(ws: Worksheet, *, header_row: int) -> int:
    last_non_empty_col = 0
    for col_idx in range(1, ws.max_column + 1):
        header_value = ws.cell(row=header_row, column=col_idx).value
        if isinstance(header_value, str) and header_value.strip():
            last_non_empty_col = col_idx
    return max(1, last_non_empty_col)


def _can_add_table(ws: Worksheet, *, header_row: int, max_col: int) -> bool:
    if max_col < 1:
        return False
    seen: set[str] = set()
    for col_idx in range(1, max_col + 1):
        header_value = ws.cell(row=header_row, column=col_idx).value
        if not isinstance(header_value, str):
            return False
        normalized = header_value.strip()
        if not normalized:
            return False
        key = normalized.casefold()
        if key in seen:
            return False
        seen.add(key)
    return True


def _add_interactive_table(
    ws: Worksheet,
    *,
    header_row: int,
    data_end_row: int,
    max_col: int,
) -> bool:
    if data_end_row <= header_row:
        return False
    if not _can_add_table(ws, header_row=header_row, max_col=max_col):
        return False
    ref = _table_range_ref(ws, header_row=header_row, data_end_row=data_end_row, max_col=max_col)
    table_name = _build_unique_table_name(ws, seed=f"{ws.title}_{header_row}")
    try:
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
        return True
    except ValueError:
        # Fail-safe: export should still succeed even if a table cannot be created.
        return False


def _apply_conditional_formatting(
    ws: Worksheet,
    *,
    header_row: int,
    data_start_row: int,
    data_end_row: int,
    header_map: dict[str, int],
    status_col_name: str,
    flags_col_name: str,
) -> None:
    if data_end_row < data_start_row:
        return

    for header_name, col_idx in header_map.items():
        if "(dk)" not in header_name:
            continue
        col_letter = get_column_letter(col_idx)
        col_range = f"{col_letter}{data_start_row}:{col_letter}{data_end_row}"
        ws.conditional_formatting.add(
            col_range,
            DataBarRule(
                start_type="num",
                start_value=0,
                end_type="max",
                end_value=0,
                color="4F81BD",
                showValue=True,
            ),
        )

    overtime_col = header_map.get("Yasal Fazla Mesai (dk)")
    if overtime_col:
        overtime_letter = get_column_letter(overtime_col)
        overtime_range = f"{overtime_letter}{data_start_row}:{overtime_letter}{data_end_row}"
        ws.conditional_formatting.add(
            overtime_range,
            ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="E6F4EA",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFF3CD",
                end_type="max",
                end_value=0,
                end_color="FDE2E4",
            ),
        )

    status_col = header_map.get(status_col_name) if status_col_name else None
    if status_col:
        status_letter = get_column_letter(status_col)
        status_range = f"{status_letter}{data_start_row}:{status_letter}{data_end_row}"
        ws.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[f'UPPER({status_letter}{data_start_row})="INCOMPLETE"'],
                fill=ALERT_FILL,
            ),
        )
        ws.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[f'OR(UPPER({status_letter}{data_start_row})="LEAVE",UPPER({status_letter}{data_start_row})="OFF")'],
                fill=WARNING_FILL,
            ),
        )
        ws.conditional_formatting.add(
            status_range,
            FormulaRule(
                formula=[f'OR(UPPER({status_letter}{data_start_row})="OK",UPPER({status_letter}{data_start_row})="FINISHED")'],
                fill=SUCCESS_FILL,
            ),
        )

    flags_col = header_map.get(flags_col_name) if flags_col_name else None
    if flags_col:
        flags_letter = get_column_letter(flags_col)
        flags_range = f"{flags_letter}{data_start_row}:{flags_letter}{data_end_row}"
        ws.conditional_formatting.add(
            flags_range,
            FormulaRule(
                formula=[f'AND({flags_letter}{data_start_row}<>"",{flags_letter}{data_start_row}<>"-")'],
                fill=ALERT_FILL,
                font=Font(bold=True, color="9F1239"),
            ),
        )


def _work_rule_minutes(db: Session, department_id: int | None, cache: dict[int | None, tuple[int, int]]) -> tuple[int, int]:
    if department_id in cache:
        return cache[department_id]

    if department_id is None:
        cache[department_id] = (DEFAULT_DAILY_MINUTES_PLANNED, DEFAULT_BREAK_MINUTES)
        return cache[department_id]

    rule = db.scalar(select(WorkRule).where(WorkRule.department_id == department_id))
    if rule is None:
        cache[department_id] = (DEFAULT_DAILY_MINUTES_PLANNED, DEFAULT_BREAK_MINUTES)
        return cache[department_id]

    cache[department_id] = (rule.daily_minutes_planned, rule.break_minutes)
    return cache[department_id]


def _gross_minutes(check_in: datetime | None, check_out: datetime | None) -> int:
    if check_in is None or check_out is None:
        return 0
    return max(0, int((check_out - check_in).total_seconds() // 60))


def _time_range_label(check_in: datetime | None, check_out: datetime | None) -> str:
    if check_in is None or check_out is None:
        return "-"
    in_ts = _to_excel_datetime(check_in)
    out_ts = _to_excel_datetime(check_out)
    if in_ts is None or out_ts is None:
        return "-"
    return f"{in_ts:%H:%M} - {out_ts:%H:%M}"


def _append_summary_area(
    ws: Worksheet,
    *,
    report: MonthlyEmployeeResponse,
    total_extra_work_minutes: int = 0,
    total_legal_overtime_minutes: int | None = None,
) -> tuple[int, int]:
    if total_legal_overtime_minutes is None:
        total_legal_overtime_minutes = report.totals.legal_overtime_minutes

    worked_days = 0
    sunday_worked_days = 0
    special_worked_days = 0
    sunday_worked_minutes = 0
    special_worked_minutes = 0
    for day in report.days:
        day_flags = list(day.flags)
        day_type = _day_type_label(day_date=day.date, leave_type=day.leave_type, flags=day_flags)
        worked_flag = _was_worked_day(
            worked_minutes=day.worked_minutes,
            check_in=day.check_in,
            check_out=day.check_out,
        )
        if worked_flag:
            worked_days += 1
            if day_type == "Pazar":
                sunday_worked_days += 1
                sunday_worked_minutes += day.worked_minutes
            if day_type == "Özel Gün":
                special_worked_days += 1
                special_worked_minutes += day.worked_minutes

    _append_spacer_row(ws)
    summary_start = ws.max_row + 1
    ws.append(["\u00d6zet", "De\u011fer"])
    ws.append(["Toplam Net S\u00fcre", _minutes_to_hhmm(report.totals.worked_minutes)])
    ws.append(["Toplam Plan \u00dcst\u00fc S\u00fcre", _minutes_to_hhmm(report.totals.plan_overtime_minutes)])
    ws.append(["Toplam Fazla S\u00fcrelerle \u00c7al\u0131\u015fma", _minutes_to_hhmm(total_extra_work_minutes)])
    ws.append(["Toplam Yasal Fazla Mesai", _minutes_to_hhmm(total_legal_overtime_minutes)])
    ws.append(["Eksik G\u00fcn", report.totals.incomplete_days])
    ws.append(["Çalışılan Gün", worked_days])
    ws.append(["Pazar Çalışılan Gün", sunday_worked_days])
    ws.append(["Özel Gün Çalışılan Gün", special_worked_days])
    ws.append(["Pazar Net Süre", _minutes_to_hhmm(sunday_worked_minutes)])
    ws.append(["Özel Gün Net Süre", _minutes_to_hhmm(special_worked_minutes)])
    ws.append(["Y\u0131ll\u0131k Fazla Mesai Kullan\u0131m\u0131", _minutes_to_hhmm(report.annual_overtime_used_minutes)])
    _style_header(ws, summary_start)

    summary_end = ws.max_row
    for row_idx in range(summary_start + 1, summary_end + 1):
        label_cell = ws.cell(row=row_idx, column=1)
        value_cell = ws.cell(row=row_idx, column=2)
        label_cell.fill = SUMMARY_FILL
        label_cell.border = THIN_BORDER
        label_cell.font = BOLD_FONT
        value_cell.fill = PatternFill(fill_type="solid", fgColor="F8FBFF")
        value_cell.border = THIN_BORDER
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        if _is_hhmm_value(value_cell.value):
            value_cell.font = Font(bold=True, color="0B4F73")
    return summary_start, summary_end


def _append_employee_daily_sheet(
    ws: Worksheet,
    *,
    employee_name: str,
    department_name: str | None,
    report: MonthlyEmployeeResponse,
    contract_weekly_minutes: int | None = None,
) -> None:
    _merge_title(ws, 1, "PUANTAJ AYLIK RAPORU", end_col=len(DAILY_HEADERS))
    ws.append(["\u00c7al\u0131\u015fan", employee_name])
    ws.append(["Departman", department_name or "-"])
    ws.append(["D\u00f6nem", f"{report.year}-{report.month:02d}"])
    ws.append(["Rapor \u00dcretim (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")])
    _append_spacer_row(ws)
    _style_metadata_rows(ws, start_row=2, end_row=5)

    header_row = ws.max_row + 1
    ws.append(DAILY_HEADERS)
    _style_header(ws, header_row)

    total_extra_work_minutes = report.totals.legal_extra_work_minutes
    total_legal_overtime_minutes = report.totals.legal_overtime_minutes
    daily_legal_breakdown = _build_daily_legal_breakdown(
        report,
        contract_weekly_minutes=contract_weekly_minutes,
    )
    data_start_row = header_row + 1

    for day in report.days:
        check_in = _to_excel_datetime(day.check_in)
        check_out = _to_excel_datetime(day.check_out)
        gross_minutes = _gross_minutes(day.check_in, day.check_out)
        break_minutes = max(0, gross_minutes - day.worked_minutes)
        day_flags = list(day.flags)
        day_type = _day_type_label(day_date=day.date, leave_type=day.leave_type, flags=day_flags)
        worked_flag = _was_worked_day(
            worked_minutes=day.worked_minutes,
            check_in=day.check_in,
            check_out=day.check_out,
        )
        plan_overtime_minutes = max(0, int(day.plan_overtime_minutes))
        extra_work_minutes, legal_overtime_minutes = daily_legal_breakdown.get(
            day.date,
            (
                max(0, int(day.legal_extra_work_minutes)),
                max(0, int(day.legal_overtime_minutes)),
            ),
        )
        ws.append(
            [
                day.date,
                check_in,
                check_out,
                _time_range_label(day.check_in, day.check_out),
                _minutes_to_excel_duration(gross_minutes),
                _minutes_to_excel_duration(break_minutes),
                _minutes_to_excel_duration(day.worked_minutes),
                day.worked_minutes,
                _minutes_to_excel_duration(plan_overtime_minutes),
                plan_overtime_minutes,
                _minutes_to_excel_duration(extra_work_minutes),
                _minutes_to_excel_duration(legal_overtime_minutes),
                legal_overtime_minutes,
                day_type,
                "Evet" if worked_flag else "Hayır",
                day.status,
                ", ".join(day_flags) if day_flags else "-",
            ]
        )

    data_end_row = ws.max_row
    _style_table_region(
        ws,
        header_row=header_row,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
    )

    _append_summary_area(
        ws,
        report=report,
        total_extra_work_minutes=total_extra_work_minutes,
        total_legal_overtime_minutes=total_legal_overtime_minutes,
    )

    for row in ws.iter_rows(min_row=data_start_row, max_row=data_end_row):
        if isinstance(row[0].value, datetime) or isinstance(row[0].value, date):
            row[0].number_format = "yyyy-mm-dd"
        if row[1].value is not None:
            row[1].number_format = "hh:mm"
        if row[2].value is not None:
            row[2].number_format = "hh:mm"
    _apply_duration_formats(
        ws,
        header_row=header_row,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        header_names=[
            "Çalışma Süresi",
            "Mola",
            "Net Süre",
            "Plan Üstü Süre",
            "Fazla Sürelerle Çalışma",
            "Yasal Fazla Mesai",
        ],
    )

    _auto_width(ws)
    _finalize_visual_scope(ws, used_max_col=len(DAILY_HEADERS))
    _apply_print_layout(ws, header_row=header_row)


def _build_employee_export(
    db: Session,
    wb: Workbook,
    *,
    employee_id: int,
    year: int,
    month: int,
) -> None:
    employee = db.get(Employee, employee_id)
    if employee is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Employee not found")

    department_name = employee.department.name if employee.department is not None else None
    report = calculate_employee_monthly(db, employee_id=employee_id, year=year, month=month)
    daily_legal_breakdown = _build_daily_legal_breakdown(
        report,
        contract_weekly_minutes=employee.contract_weekly_minutes,
    )
    daily_rows: list[dict[str, object]] = []
    for day in report.days:
        day_type = _day_type_label(
            day_date=day.date,
            leave_type=day.leave_type,
            flags=list(day.flags),
        )
        worked_flag = _was_worked_day(
            worked_minutes=day.worked_minutes,
            check_in=day.check_in,
            check_out=day.check_out,
        )
        daily_rows.append(
            {
                "date": day.date,
                "employee_id": employee.id,
                "employee_name": employee.full_name,
                "department_name": department_name or "-",
                "worked_minutes": day.worked_minutes,
                "plan_overtime_minutes": day.plan_overtime_minutes,
                "extra_work_minutes": daily_legal_breakdown.get(
                    day.date,
                    (day.legal_extra_work_minutes, day.legal_overtime_minutes),
                )[0],
                "overtime_minutes": daily_legal_breakdown.get(
                    day.date,
                    (day.legal_extra_work_minutes, day.legal_overtime_minutes),
                )[1],
                "status": day.status,
                "day_type": day_type,
                "worked_flag": worked_flag,
            }
        )

    summary_rows = [
        _build_summary_row_from_daily_facts(
            employee_id=employee.id,
            employee_name=employee.full_name,
            department_name=department_name or "-",
            rows=daily_rows,
        )
    ]
    summary_rows[0]["annual_overtime_minutes"] = report.annual_overtime_used_minutes

    ws_daily = wb.create_sheet(_safe_sheet_title(f"DAILY_{employee.id}", "DAILY_FACT"))
    _append_employee_daily_sheet(
        ws_daily,
        employee_name=employee.full_name,
        department_name=department_name,
        report=report,
        contract_weekly_minutes=employee.contract_weekly_minutes,
    )

    _build_dashboard_sheet(
        wb.active,
        title="Çalışan Aylık Özeti",
        summary_rows=summary_rows,
        employee_sheet_map={employee.id: ws_daily.title},
    )
    _build_summary_sheet(
        wb.create_sheet("SUMMARY_EMPLOYEE"),
        title="Çalışan Aylık Özeti",
        rows=summary_rows,
    )
    _build_department_summary_sheet(
        wb.create_sheet("SUMMARY_DEPARTMENT"),
        title="Çalışan Aylık Özeti",
        rows=_group_summary_rows_by_department(summary_rows),
    )


def _build_summary_sheet(
    ws: Worksheet,
    *,
    title: str,
    rows: list[dict[str, object]],
) -> None:
    _merge_title(ws, 1, f"{title} - PUANTAJ OZET", end_col=14)
    ws.append(["Rapor \u00dcretim (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Kay\u0131t Say\u0131s\u0131", len(rows)])
    _append_spacer_row(ws)
    _style_metadata_rows(ws, start_row=2, end_row=3)

    header_row = ws.max_row + 1
    ws.append(
        [
            "\u00c7al\u0131\u015fan ID",
            "\u00c7al\u0131\u015fan Ad\u0131",
            "Departman",
            "Toplam Net S\u00fcre",
            "Toplam Plan \u00dcst\u00fc S\u00fcre",
            "Toplam Fazla S\u00fcrelerle \u00c7al\u0131\u015fma",
            "Toplam Yasal Fazla Mesai",
            "Eksik G\u00fcn",
            "Çalışılan Gün",
            "Pazar Çalışılan Gün",
            "Özel Gün Çalışılan Gün",
            "Pazar Net Süre",
            "Özel Gün Net Süre",
            "Y\u0131ll\u0131k Fazla Mesai Kullan\u0131m\u0131",
        ]
    )
    _style_header(ws, header_row)

    total_worked = 0
    total_plan_overtime = 0
    total_extra = 0
    total_overtime = 0
    total_incomplete = 0
    total_worked_days = 0
    total_sunday_days = 0
    total_special_days = 0
    total_sunday_minutes = 0
    total_special_minutes = 0
    total_annual = 0
    data_start_row = header_row + 1
    for row in rows:
        worked = int(row["worked_minutes"])
        plan_overtime = int(row.get("plan_overtime_minutes", 0))
        extra = int(row["extra_work_minutes"])
        overtime = int(row["overtime_minutes"])
        incomplete = int(row["incomplete_days"])
        worked_days = int(row.get("worked_days", 0))
        sunday_days = int(row.get("sunday_worked_days", 0))
        special_days = int(row.get("special_worked_days", 0))
        sunday_minutes = int(row.get("sunday_worked_minutes", 0))
        special_minutes = int(row.get("special_worked_minutes", 0))
        annual = int(row["annual_overtime_minutes"])

        total_worked += worked
        total_plan_overtime += plan_overtime
        total_extra += extra
        total_overtime += overtime
        total_incomplete += incomplete
        total_worked_days += worked_days
        total_sunday_days += sunday_days
        total_special_days += special_days
        total_sunday_minutes += sunday_minutes
        total_special_minutes += special_minutes
        total_annual += annual

        ws.append(
            [
                row["employee_id"],
                row["employee_name"],
                row["department_name"],
                _minutes_to_excel_duration(worked),
                _minutes_to_excel_duration(plan_overtime),
                _minutes_to_excel_duration(extra),
                _minutes_to_excel_duration(overtime),
                incomplete,
                worked_days,
                sunday_days,
                special_days,
                _minutes_to_excel_duration(sunday_minutes),
                _minutes_to_excel_duration(special_minutes),
                _minutes_to_hhmm(annual),
            ]
        )

    data_end_row = ws.max_row
    _style_table_region(
        ws,
        header_row=header_row,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        status_col_name="",
        flags_col_name="",
    )

    total_row = ws.max_row + 1
    ws.append(
        [
            "Toplam",
            "",
            "",
            _minutes_to_excel_duration(total_worked),
            _minutes_to_excel_duration(total_plan_overtime),
            _minutes_to_excel_duration(total_extra),
            _minutes_to_excel_duration(total_overtime),
            total_incomplete,
            total_worked_days,
            total_sunday_days,
            total_special_days,
            _minutes_to_excel_duration(total_sunday_minutes),
            _minutes_to_excel_duration(total_special_minutes),
            _minutes_to_hhmm(total_annual),
        ]
    )
    for cell in ws[total_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    _apply_duration_formats(
        ws,
        header_row=header_row,
        data_start_row=data_start_row,
        data_end_row=total_row,
        header_names=[
            "Toplam Net Süre",
            "Toplam Plan Üstü Süre",
            "Toplam Fazla Sürelerle Çalışma",
            "Toplam Yasal Fazla Mesai",
            "Pazar Net Süre",
            "Özel Gün Net Süre",
        ],
    )
    _auto_width(ws)
    _finalize_visual_scope(ws, used_max_col=14)
    _apply_print_layout(ws, header_row=header_row)


def _build_department_summary_sheet(
    ws: Worksheet,
    *,
    title: str,
    rows: list[dict[str, object]],
) -> None:
    _merge_title(ws, 1, f"{title} - DEPARTMAN KPI", end_col=9)
    ws.append(["Rapor Üretim (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Departman Sayısı", len(rows)])
    _append_spacer_row(ws)
    _style_metadata_rows(ws, start_row=2, end_row=3)

    header_row = ws.max_row + 1
    ws.append(
        [
            "Departman",
            "Çalışan Sayısı",
            "Toplam Net Süre",
            "Toplam Plan Üstü Süre",
            "Toplam Fazla Sürelerle Çalışma",
            "Toplam Yasal Fazla Mesai",
            "Eksik Gün",
            "Pazar Çalışılan Gün",
            "Özel Gün Çalışılan Gün",
        ]
    )
    _style_header(ws, header_row)

    data_start_row = header_row + 1
    total_employee_count = 0
    total_worked = 0
    total_plan_overtime = 0
    total_extra = 0
    total_overtime = 0
    total_incomplete = 0
    total_sunday_days = 0
    total_special_days = 0
    for row in rows:
        employee_count = int(row["employee_count"])
        worked_minutes = int(row["worked_minutes"])
        plan_overtime_minutes = int(row.get("plan_overtime_minutes", 0))
        extra_work_minutes = int(row["extra_work_minutes"])
        overtime_minutes = int(row["overtime_minutes"])
        incomplete_days = int(row["incomplete_days"])
        sunday_days = int(row.get("sunday_worked_days", 0))
        special_days = int(row.get("special_worked_days", 0))

        total_employee_count += employee_count
        total_worked += worked_minutes
        total_plan_overtime += plan_overtime_minutes
        total_extra += extra_work_minutes
        total_overtime += overtime_minutes
        total_incomplete += incomplete_days
        total_sunday_days += sunday_days
        total_special_days += special_days

        ws.append(
            [
                row["department_name"],
                employee_count,
                _minutes_to_excel_duration(worked_minutes),
                _minutes_to_excel_duration(plan_overtime_minutes),
                _minutes_to_excel_duration(extra_work_minutes),
                _minutes_to_excel_duration(overtime_minutes),
                incomplete_days,
                sunday_days,
                special_days,
            ]
        )

    data_end_row = ws.max_row
    _style_table_region(
        ws,
        header_row=header_row,
        data_start_row=data_start_row,
        data_end_row=data_end_row,
        status_col_name="",
        flags_col_name="",
    )

    total_row = ws.max_row + 1
    ws.append(
        [
            "Toplam",
            total_employee_count,
            _minutes_to_excel_duration(total_worked),
            _minutes_to_excel_duration(total_plan_overtime),
            _minutes_to_excel_duration(total_extra),
            _minutes_to_excel_duration(total_overtime),
            total_incomplete,
            total_sunday_days,
            total_special_days,
        ]
    )
    for cell in ws[total_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    _apply_duration_formats(
        ws,
        header_row=header_row,
        data_start_row=data_start_row,
        data_end_row=total_row,
        header_names=[
            "Toplam Net Süre",
            "Toplam Plan Üstü Süre",
            "Toplam Fazla Sürelerle Çalışma",
            "Toplam Yasal Fazla Mesai",
        ],
    )
    _auto_width(ws)
    _finalize_visual_scope(ws, used_max_col=9)
    _apply_print_layout(ws, header_row=header_row)


def _build_dashboard_sheet(
    ws: Worksheet,
    *,
    title: str,
    summary_rows: list[dict[str, object]],
    employee_sheet_map: dict[int, str] | None = None,
) -> None:
    ws.title = _safe_sheet_title("DASHBOARD", "Dashboard")
    _merge_title(ws, 1, f"{title} - YÖNETİM DASHBOARD", end_col=4)
    ws.append(["Rapor Üretim (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")])
    _append_spacer_row(ws)
    _style_metadata_rows(ws, start_row=2, end_row=2)

    total_employee = len(summary_rows)
    total_worked = sum(int(row["worked_minutes"]) for row in summary_rows)
    total_plan_overtime = sum(int(row.get("plan_overtime_minutes", 0)) for row in summary_rows)
    total_extra = sum(int(row["extra_work_minutes"]) for row in summary_rows)
    total_overtime = sum(int(row["overtime_minutes"]) for row in summary_rows)
    total_incomplete = sum(int(row["incomplete_days"]) for row in summary_rows)

    kpi_row = ws.max_row + 1
    ws.append(["KPI", "Değer"])
    _style_header(ws, kpi_row)
    kpi_rows = [
        ("Toplam Çalışan", total_employee, False),
        ("Toplam Net Süre", _minutes_to_excel_duration(total_worked), True),
        ("Toplam Plan Üstü Süre", _minutes_to_excel_duration(total_plan_overtime), True),
        ("Toplam Fazla Sürelerle Çalışma", _minutes_to_excel_duration(total_extra), True),
        ("Toplam Yasal Fazla Mesai", _minutes_to_excel_duration(total_overtime), True),
        ("Toplam Eksik Gün", total_incomplete, False),
    ]
    for label, value, _ in kpi_rows:
        ws.append([label, value])
    _style_metadata_rows(ws, start_row=kpi_row + 1, end_row=ws.max_row)
    for offset, (_, _, is_duration) in enumerate(kpi_rows, start=1):
        value_cell = ws.cell(row=kpi_row + offset, column=2)
        value_cell.number_format = DURATION_NUMBER_FORMAT if is_duration else "0"

    _append_spacer_row(ws)
    overtime_header_row = ws.max_row + 1
    ws.append(["Sıra", "Çalışan Navigasyon", "Yasal Fazla Mesai (dk)", "Yasal Fazla Mesai [h:mm]"])
    _style_header(ws, overtime_header_row)
    top_rows = sorted(summary_rows, key=lambda item: int(item["overtime_minutes"]), reverse=True)
    overtime_data_start = ws.max_row + 1
    for idx, row in enumerate(top_rows, start=1):
        overtime_minutes = int(row["overtime_minutes"])
        ws.append(
            [
                idx,
                f"#{row['employee_id']} - {row['employee_name']}",
                overtime_minutes,
                _minutes_to_excel_duration(overtime_minutes),
            ]
        )
    overtime_data_end = ws.max_row
    _style_table_region(
        ws,
        header_row=overtime_header_row,
        data_start_row=overtime_data_start,
        data_end_row=overtime_data_end,
        status_col_name="",
        flags_col_name="",
    )
    _apply_duration_formats(
        ws,
        header_row=overtime_header_row,
        data_start_row=overtime_data_start,
        data_end_row=overtime_data_end,
        header_names=["Yasal Fazla Mesai [h:mm]"],
    )
    if employee_sheet_map:
        for row_idx in range(overtime_data_start, overtime_data_end + 1):
            name_cell = ws.cell(row=row_idx, column=2)
            value = str(name_cell.value or "")
            if not value.startswith("#") or " - " not in value:
                continue
            employee_id_token = value.split(" - ", 1)[0].lstrip("#")
            if not employee_id_token.isdigit():
                continue
            employee_sheet_title = employee_sheet_map.get(int(employee_id_token))
            if employee_sheet_title:
                _set_internal_sheet_link(name_cell, employee_sheet_title)

    _append_spacer_row(ws)
    dep_header = ws.max_row + 1
    ws.append(["Departman", "Net Süre (dk)", "Plan Üstü Süre (dk)", "Yasal Fazla Mesai (dk)"])
    _style_header(ws, dep_header)
    department_rows = _group_summary_rows_by_department(summary_rows)
    dep_data_start = ws.max_row + 1
    for row in department_rows:
        ws.append(
            [
                row["department_name"],
                int(row["worked_minutes"]),
                int(row.get("plan_overtime_minutes", 0)),
                int(row["overtime_minutes"]),
            ]
        )
    dep_data_end = ws.max_row
    _style_table_region(
        ws,
        header_row=dep_header,
        data_start_row=dep_data_start,
        data_end_row=dep_data_end,
        status_col_name="",
        flags_col_name="",
    )

    _auto_width(ws)
    _finalize_visual_scope(ws, used_max_col=4)
    _apply_print_layout(ws)


def _build_department_or_all_export(
    db: Session,
    wb: Workbook,
    *,
    year: int,
    month: int,
    department_id: int | None,
    include_daily_sheet: bool,
    include_inactive: bool,
) -> None:
    employee_stmt = select(Employee).order_by(Employee.id.asc())
    if not include_inactive:
        employee_stmt = employee_stmt.where(Employee.is_active.is_(True))
    if department_id is not None:
        employee_stmt = employee_stmt.where(Employee.department_id == department_id)
    employees = list(db.scalars(employee_stmt).all())

    ws_dashboard = wb.active
    summary_rows: list[dict[str, object]] = []
    employee_sheet_map: dict[int, str] = {}
    for employee in employees:
        report = calculate_employee_monthly(db, employee_id=employee.id, year=year, month=month)
        department_name = employee.department.name if employee.department is not None else "-"
        worked_days = 0
        sunday_worked_days = 0
        special_worked_days = 0
        sunday_worked_minutes = 0
        special_worked_minutes = 0
        for day in report.days:
            day_type = _day_type_label(
                day_date=day.date,
                leave_type=day.leave_type,
                flags=list(day.flags),
            )
            worked_flag = _was_worked_day(
                worked_minutes=day.worked_minutes,
                check_in=day.check_in,
                check_out=day.check_out,
            )
            if worked_flag:
                worked_days += 1
                if day_type == "Pazar":
                    sunday_worked_days += 1
                    sunday_worked_minutes += day.worked_minutes
                if day_type == "Özel Gün":
                    special_worked_days += 1
                    special_worked_minutes += day.worked_minutes
        summary_rows.append(
            {
                "employee_id": employee.id,
                "employee_name": employee.full_name,
                "department_name": department_name,
                "worked_minutes": report.totals.worked_minutes,
                "plan_overtime_minutes": report.totals.plan_overtime_minutes,
                "extra_work_minutes": report.totals.legal_extra_work_minutes,
                "overtime_minutes": report.totals.legal_overtime_minutes,
                "incomplete_days": report.totals.incomplete_days,
                "annual_overtime_minutes": report.annual_overtime_used_minutes,
                "worked_days": worked_days,
                "sunday_worked_days": sunday_worked_days,
                "special_worked_days": special_worked_days,
                "sunday_worked_minutes": sunday_worked_minutes,
                "special_worked_minutes": special_worked_minutes,
            }
        )
        if include_daily_sheet:
            ws_daily = wb.create_sheet(
                _safe_sheet_title(f"{employee.full_name} ({employee.id})", f"Emp{employee.id}")
            )
            employee_sheet_map[employee.id] = ws_daily.title
            _append_employee_daily_sheet(
                ws_daily,
                employee_name=employee.full_name,
                department_name=department_name,
                report=report,
                contract_weekly_minutes=employee.contract_weekly_minutes,
            )

    summary_title = "Departman Özeti" if department_id is not None else "Tüm Çalışanlar Özeti"
    _build_dashboard_sheet(
        ws_dashboard,
        title=summary_title,
        summary_rows=summary_rows,
        employee_sheet_map=employee_sheet_map if employee_sheet_map else None,
    )
    _build_summary_sheet(wb.create_sheet("SUMMARY_EMPLOYEE"), title=summary_title, rows=summary_rows)
    _build_department_summary_sheet(
        wb.create_sheet("SUMMARY_DEPARTMENT"),
        title=summary_title,
        rows=_group_summary_rows_by_department(summary_rows),
    )


def _build_date_range_export(
    db: Session,
    wb: Workbook,
    *,
    start_date: date,
    end_date: date,
    employee_id: int | None,
    department_id: int | None,
) -> None:
    if end_date < start_date:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="end_date must be >= start_date",
        )

    start_dt, end_dt = _date_range_bounds_utc(start_date, end_date)

    stmt = (
        select(AttendanceEvent, Employee, Department)
        .join(Employee, AttendanceEvent.employee_id == Employee.id)
        .outerjoin(Department, Employee.department_id == Department.id)
        .where(
            AttendanceEvent.ts_utc >= start_dt,
            AttendanceEvent.ts_utc < end_dt,
            AttendanceEvent.deleted_at.is_(None),
        )
        .order_by(AttendanceEvent.ts_utc.asc(), AttendanceEvent.id.asc())
    )
    if employee_id is not None:
        stmt = stmt.where(AttendanceEvent.employee_id == employee_id)
    if department_id is not None:
        stmt = stmt.where(Employee.department_id == department_id)

    rows = list(db.execute(stmt).all())

    ws_events = wb.active
    ws_events.title = "RAW_EVENTS"
    _merge_title(ws_events, 1, "PUANTAJ RAW EVENTS RAPORU", end_col=11)
    ws_events.append(["Tarih Aral\u0131\u011f\u0131", f"{start_date.isoformat()} - {end_date.isoformat()}"])
    ws_events.append(["Filtre - \u00c7al\u0131\u015fan ID", employee_id if employee_id is not None else "T\u00fcm\u00fc"])
    ws_events.append(["Filtre - Departman ID", department_id if department_id is not None else "T\u00fcm\u00fc"])
    ws_events.append(["Kay\u0131t Say\u0131s\u0131", len(rows)])
    ws_events.append([])
    _style_metadata_rows(ws_events, start_row=2, end_row=5)

    events_header_row = ws_events.max_row + 1
    ws_events.append(
        [
            "Event ID",
            "\u00c7al\u0131\u015fan ID",
            "\u00c7al\u0131\u015fan Ad\u0131",
            "Departman",
            "Tip",
            "Zaman (UTC)",
            "Konum Durumu",
            "Enlem",
            "Boylam",
            "Doğruluk (m)",
            "Bayraklar",
        ]
    )
    _style_header(ws_events, events_header_row)

    grouped: dict[tuple[int, date], dict[str, object]] = defaultdict(
        lambda: {
            "employee_name": "",
            "department_name": "-",
            "department_id": None,
            "employee_ref": None,
            "shift_name": "-",
            "ins": [],
            "outs": [],
            "flags": set(),
        }
    )
    for event, employee, department in rows:
        flags = ", ".join([k for k, v in (event.flags or {}).items() if v is True]) or "-"
        ws_events.append(
            [
                event.id,
                event.employee_id,
                employee.full_name,
                department.name if department is not None else "-",
                event.type.value,
                _to_excel_datetime(event.ts_utc),
                event.location_status.value,
                event.lat,
                event.lon,
                event.accuracy_m,
                flags,
            ]
        )

        key = (event.employee_id, _local_day_from_ts(event.ts_utc))
        grouped[key]["employee_name"] = employee.full_name
        grouped[key]["department_name"] = department.name if department is not None else "-"
        grouped[key]["department_id"] = employee.department_id
        grouped[key]["employee_ref"] = employee
        shift_name = (event.flags or {}).get("SHIFT_NAME")
        if isinstance(shift_name, str) and shift_name.strip():
            grouped[key]["shift_name"] = shift_name.strip()
        if event.type == AttendanceType.IN:
            grouped[key]["ins"].append(event)
        else:
            grouped[key]["outs"].append(event)
        for flag_name, flag_value in (event.flags or {}).items():
            if isinstance(flag_value, bool) and flag_value:
                grouped[key]["flags"].add(flag_name)

    events_data_start = events_header_row + 1
    events_data_end = ws_events.max_row
    _style_table_region(
        ws_events,
        header_row=events_header_row,
        data_start_row=events_data_start,
        data_end_row=events_data_end,
        status_col_name="Konum Durumu",
        flags_col_name="Bayraklar",
    )

    for row in ws_events.iter_rows(min_row=events_data_start, max_row=events_data_end):
        if row[5].value is not None:
            row[5].number_format = "yyyy-mm-dd hh:mm"
    _auto_width(ws_events)
    _finalize_visual_scope(ws_events, used_max_col=11)
    _apply_print_layout(ws_events, header_row=events_header_row)

    ws_daily = wb.create_sheet("DAILY_FACT")
    _merge_title(ws_daily, 1, "PUANTAJ DAILY FACT RAPORU", end_col=21)
    ws_daily.append(["Tarih Aral\u0131\u011f\u0131", f"{start_date.isoformat()} - {end_date.isoformat()}"])
    ws_daily.append(["Filtre - \u00c7al\u0131\u015fan ID", employee_id if employee_id is not None else "T\u00fcm\u00fc"])
    ws_daily.append(["Filtre - Departman ID", department_id if department_id is not None else "T\u00fcm\u00fc"])
    ws_daily.append([])
    _style_metadata_rows(ws_daily, start_row=2, end_row=4)

    daily_header_row = ws_daily.max_row + 1
    ws_daily.append(
        [
            "Tarih",
            "Çalışan ID",
            "Çalışan Adı",
            "Departman",
            "Vardiya",
            "Giriş",
            "Çıkış",
            "Saat Aralığı",
            "Çalışma Süresi",
            "Mola",
            "Net Süre",
            "Net Süre (dk)",
            "Plan Üstü Süre",
            "Plan Üstü Süre (dk)",
            "Fazla Sürelerle Çalışma",
            "Yasal Fazla Mesai",
            "Yasal Fazla Mesai (dk)",
            "Gün Tipi",
            "Çalışıldı mı",
            "Durum",
            "Bayraklar",
        ]
    )
    _style_header(ws_daily, daily_header_row)

    work_rule_cache: dict[int | None, tuple[int, int]] = {}
    monthly_report_cache: dict[tuple[int, int, int], MonthlyEmployeeResponse] = {}
    monthly_day_lookup_cache: dict[tuple[int, int, int], dict[date, object]] = {}
    monthly_legal_lookup_cache: dict[tuple[int, int, int], dict[date, tuple[int, int]]] = {}

    def _resolve_monthly_context(
        employee_row: Employee,
        day_value: date,
    ) -> tuple[dict[date, object], dict[date, tuple[int, int]]]:
        cache_key = (employee_row.id, day_value.year, day_value.month)
        if cache_key not in monthly_report_cache:
            report = calculate_employee_monthly(
                db,
                employee_id=employee_row.id,
                year=day_value.year,
                month=day_value.month,
            )
            monthly_report_cache[cache_key] = report
            monthly_day_lookup_cache[cache_key] = {item.date: item for item in report.days}
            monthly_legal_lookup_cache[cache_key] = _build_daily_legal_breakdown(
                report,
                contract_weekly_minutes=employee_row.contract_weekly_minutes,
            )
        return monthly_day_lookup_cache[cache_key], monthly_legal_lookup_cache[cache_key]

    total_worked_minutes = 0
    total_plan_overtime_minutes = 0
    total_extra_work_minutes = 0
    total_overtime_minutes = 0
    daily_fact_rows: list[dict[str, object]] = []
    for (employee_id_value, day_date), bucket in sorted(grouped.items(), key=lambda item: (item[0][1], item[0][0])):
        ins = bucket["ins"]
        outs = bucket["outs"]
        first_in = ins[0].ts_utc if ins else None
        last_out = outs[-1].ts_utc if outs else None
        employee_ref = bucket.get("employee_ref")
        if not isinstance(employee_ref, Employee):
            continue

        planned_minutes, break_minutes = _work_rule_minutes(
            db,
            bucket["department_id"],
            work_rule_cache,
        )
        day_status, worked_minutes, calculated_plan_overtime_minutes = calculate_work_and_overtime(
            first_in_ts=first_in,
            last_out_ts=last_out,
            planned_minutes=planned_minutes,
            break_minutes=break_minutes,
        )
        plan_overtime_minutes = max(0, int(calculated_plan_overtime_minutes))
        monthly_day_lookup, monthly_legal_lookup = _resolve_monthly_context(employee_ref, day_date)
        monthly_day = monthly_day_lookup.get(day_date)
        leave_type_value: object = None
        if monthly_day is not None:
            day_status = str(getattr(monthly_day, "status", day_status))
            worked_minutes = max(0, int(getattr(monthly_day, "worked_minutes", worked_minutes)))
            plan_overtime_minutes = max(
                0,
                int(
                    getattr(
                        monthly_day,
                        "plan_overtime_minutes",
                        getattr(monthly_day, "overtime_minutes", plan_overtime_minutes),
                    )
                ),
            )
            leave_type_value = getattr(monthly_day, "leave_type", None)
        extra_work_minutes, legal_overtime_minutes = monthly_legal_lookup.get(
            day_date,
            (0, max(0, int(plan_overtime_minutes))),
        )
        gross_minutes = _gross_minutes(first_in, last_out)
        break_deducted = max(0, gross_minutes - worked_minutes)
        day_type = _day_type_label(
            day_date=day_date,
            leave_type=leave_type_value,
            flags=list(bucket["flags"]),
        )
        worked_flag = _was_worked_day(
            worked_minutes=worked_minutes,
            check_in=first_in,
            check_out=last_out,
        )
        total_worked_minutes += worked_minutes
        total_plan_overtime_minutes += plan_overtime_minutes
        total_extra_work_minutes += extra_work_minutes
        total_overtime_minutes += legal_overtime_minutes
        flag_names = sorted(bucket["flags"])
        if first_in is None:
            flag_names.append("MISSING_IN")
        if last_out is None:
            flag_names.append("MISSING_OUT")

        ws_daily.append(
            [
                day_date,
                employee_id_value,
                bucket["employee_name"],
                bucket["department_name"],
                bucket["shift_name"],
                _to_excel_datetime(first_in),
                _to_excel_datetime(last_out),
                _time_range_label(first_in, last_out),
                _minutes_to_excel_duration(gross_minutes),
                _minutes_to_excel_duration(break_deducted),
                _minutes_to_excel_duration(worked_minutes),
                worked_minutes,
                _minutes_to_excel_duration(plan_overtime_minutes),
                plan_overtime_minutes,
                _minutes_to_excel_duration(extra_work_minutes),
                _minutes_to_excel_duration(legal_overtime_minutes),
                legal_overtime_minutes,
                day_type,
                "Evet" if worked_flag else "Hayır",
                day_status,
                ", ".join(flag_names) if flag_names else "-",
            ]
        )
        daily_fact_rows.append(
            {
                "date": day_date,
                "employee_id": employee_id_value,
                "employee_name": str(bucket["employee_name"]),
                "department_name": str(bucket["department_name"]),
                "worked_minutes": worked_minutes,
                "plan_overtime_minutes": plan_overtime_minutes,
                "extra_work_minutes": extra_work_minutes,
                "overtime_minutes": legal_overtime_minutes,
                "status": day_status,
                "day_type": day_type,
                "worked_flag": worked_flag,
            }
        )

    daily_data_start = daily_header_row + 1
    daily_data_end = ws_daily.max_row
    _style_table_region(
        ws_daily,
        header_row=daily_header_row,
        data_start_row=daily_data_start,
        data_end_row=daily_data_end,
    )

    for row in ws_daily.iter_rows(min_row=daily_data_start, max_row=daily_data_end):
        if row[0].value is not None:
            row[0].number_format = "yyyy-mm-dd"
        if row[5].value is not None:
            row[5].number_format = "hh:mm"
        if row[6].value is not None:
            row[6].number_format = "hh:mm"
    _apply_duration_formats(
        ws_daily,
        header_row=daily_header_row,
        data_start_row=daily_data_start,
        data_end_row=daily_data_end,
        header_names=[
            "Çalışma Süresi",
            "Mola",
            "Net Süre",
            "Plan Üstü Süre",
            "Fazla Sürelerle Çalışma",
            "Yasal Fazla Mesai",
        ],
    )

    ws_daily.append([])
    summary_start = ws_daily.max_row + 1
    ws_daily.append(["Toplam \u00c7al\u0131\u015f\u0131lan Net S\u00fcre", _minutes_to_hhmm(total_worked_minutes)])
    ws_daily.append(["Toplam Plan \u00dcst\u00fc S\u00fcre", _minutes_to_hhmm(total_plan_overtime_minutes)])
    ws_daily.append(["Toplam Fazla S\u00fcrelerle \u00c7al\u0131\u015fma", _minutes_to_hhmm(total_extra_work_minutes)])
    ws_daily.append(["Toplam Yasal Fazla Mesai", _minutes_to_hhmm(total_overtime_minutes)])
    _style_metadata_rows(ws_daily, start_row=summary_start, end_row=ws_daily.max_row)
    _auto_width(ws_daily)
    _finalize_visual_scope(ws_daily, used_max_col=21)
    _apply_print_layout(ws_daily, header_row=daily_header_row)

    employee_daily_map: dict[tuple[int, str, str], list[dict[str, object]]] = defaultdict(list)
    for row in daily_fact_rows:
        key = (
            int(row["employee_id"]),
            str(row["employee_name"]),
            str(row["department_name"]),
        )
        employee_daily_map[key].append(row)

    summary_rows = [
        _build_summary_row_from_daily_facts(
            employee_id=employee_id_value,
            employee_name=employee_name,
            department_name=department_name,
            rows=rows_for_employee,
        )
        for (employee_id_value, employee_name, department_name), rows_for_employee in sorted(
            employee_daily_map.items(),
            key=lambda item: item[0][0],
        )
    ]

    summary_title = "Tarih Aralığı Özeti"
    _build_dashboard_sheet(
        wb.create_sheet("DASHBOARD", 0),
        title=summary_title,
        summary_rows=summary_rows,
    )
    _build_summary_sheet(
        wb.create_sheet("SUMMARY_EMPLOYEE"),
        title=summary_title,
        rows=summary_rows,
    )
    _build_department_summary_sheet(
        wb.create_sheet("SUMMARY_DEPARTMENT"),
        title=summary_title,
        rows=_group_summary_rows_by_department(summary_rows),
    )


def _sheet_purpose_text(sheet_name: str) -> str:
    upper_name = sheet_name.upper()
    if upper_name == "DASHBOARD":
        return "Yonetim KPI ve hizli durum ozeti"
    if upper_name.startswith("SUMMARY_EMPLOYEE"):
        return "Calisan bazli toplu ozet tablosu"
    if upper_name.startswith("SUMMARY_DEPARTMENT"):
        return "Departman bazli toplu ozet tablosu"
    if upper_name.startswith("RAW_EVENTS"):
        return "Ham event kayitlari (denetim)"
    if upper_name.startswith("DAILY_FACT"):
        return "Gunluk mesai gerceklik tablosu"
    if upper_name.startswith("EMP_") or upper_name.startswith("DEP_") or upper_name.startswith("DAILY_"):
        return "Detay calisma sayfasi"
    return "Rapor sayfasi"


def _apply_workbook_branding(wb: Workbook) -> None:
    for ws in wb.worksheets:
        upper_name = ws.title.upper()
        if upper_name == "README":
            ws.sheet_properties.tabColor = "FFB91C1C"
        elif upper_name == "DASHBOARD":
            ws.sheet_properties.tabColor = "FF0B4F73"
        elif upper_name.startswith("SUMMARY"):
            ws.sheet_properties.tabColor = "FF2B6A99"
        elif upper_name.startswith("RAW") or upper_name.startswith("DAILY"):
            ws.sheet_properties.tabColor = "FF475569"
        else:
            ws.sheet_properties.tabColor = "FF64748B"


def _build_readme_sheet(wb: Workbook, *, report_title: str) -> None:
    if "README" in wb.sheetnames:
        del wb["README"]
    ws = wb.create_sheet("README", 0)

    _merge_title(ws, 1, f"{report_title} - RAPOR KILAVUZU", end_col=4)
    ws.append(["Rapor Uretim (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")])
    ws.append(["Toplam Sayfa", len(wb.sheetnames) - 1])
    _append_spacer_row(ws)
    _style_metadata_rows(ws, start_row=2, end_row=3)

    guide_header = ws.max_row + 1
    ws.append(["Adim", "Ne yapmali?", "Neden onemli?"])
    _style_header(ws, guide_header)
    guide_rows = [
        ("1", "DASHBOARD sayfasindan baslayin.", "En kritik KPI metriklerini tek yerde gorursunuz."),
        ("2", "Tablolarda filtreyi acip ihtiyaciniza gore daraltin.", "Buyuk veri setlerinde hizli analiz saglar."),
        ("3", "Calisan satirlarindaki linklerle detay sayfalarina gecin.", "Kaynak veriye hizli inis yaparsiniz."),
        ("4", "SUMMARY sayfalarindan aylik/departman karsilastirmasi yapin.", "Kurumsal kararlar icin toplu bakis verir."),
    ]
    guide_start = ws.max_row + 1
    for row in guide_rows:
        ws.append(list(row))
    guide_end = ws.max_row
    _style_table_region(
        ws,
        header_row=guide_header,
        data_start_row=guide_start,
        data_end_row=guide_end,
        status_col_name="",
        flags_col_name="",
    )
    for row_idx in range(guide_start, guide_end + 1):
        ws.cell(row=row_idx, column=1).fill = README_PANEL_FILL
        ws.cell(row=row_idx, column=1).font = BOLD_FONT

    _append_spacer_row(ws)
    nav_header = ws.max_row + 1
    ws.append(["Sira", "Sayfa", "Kategori", "Aciklama"])
    _style_header(ws, nav_header)
    nav_start = ws.max_row + 1
    sheet_order = [sheet for sheet in wb.sheetnames if sheet != "README"]
    for idx, sheet_name in enumerate(sheet_order, start=1):
        ws.append([idx, sheet_name, _sheet_purpose_text(sheet_name), "Ac"])
        _set_internal_sheet_link(ws.cell(row=ws.max_row, column=2), sheet_name)
        _set_internal_sheet_link(ws.cell(row=ws.max_row, column=4), sheet_name)
    nav_end = ws.max_row
    _style_table_region(
        ws,
        header_row=nav_header,
        data_start_row=nav_start,
        data_end_row=nav_end,
        status_col_name="",
        flags_col_name="",
    )
    for row_idx in range(nav_start, nav_end + 1):
        ws.cell(row=row_idx, column=3).fill = NEUTRAL_PANEL_FILL

    _auto_width(ws)
    _finalize_visual_scope(ws, used_max_col=4)
    _apply_print_layout(ws, header_row=nav_header)


def build_puantaj_xlsx_bytes(
    db: Session,
    *,
    mode: ExportMode,
    year: int | None = None,
    month: int | None = None,
    employee_id: int | None = None,
    department_id: int | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    include_daily_sheet: bool = True,
    include_inactive: bool = False,
) -> bytes:
    wb = Workbook()

    if mode == "employee":
        if employee_id is None or year is None or month is None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="employee_id, year and month are required",
            )
        _build_employee_export(db, wb, employee_id=employee_id, year=year, month=month)
    elif mode in {"department", "all"}:
        if year is None or month is None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="year and month are required",
            )
        if mode == "department" and department_id is None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="department_id is required for department export",
            )
        _build_department_or_all_export(
            db,
            wb,
            year=year,
            month=month,
            department_id=department_id if mode == "department" else None,
            include_daily_sheet=include_daily_sheet,
            include_inactive=include_inactive,
        )
    elif mode == "date_range":
        if start_date is None or end_date is None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="start_date and end_date are required",
            )
        _build_date_range_export(
            db,
            wb,
            start_date=start_date,
            end_date=end_date,
            employee_id=employee_id,
            department_id=department_id,
        )
    else:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Invalid export mode")

    report_title_map: dict[ExportMode, str] = {
        "employee": "Calisan Aylik Raporu",
        "department": "Departman Aylik Raporu",
        "all": "Genel Aylik Rapor",
        "date_range": "Tarih Araligi Raporu",
    }
    _build_readme_sheet(wb, report_title=report_title_map.get(mode, "Puantaj Raporu"))
    _apply_workbook_branding(wb)

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _fetch_date_range_rows(
    db: Session,
    *,
    start_date: date,
    end_date: date,
    employee_id: int | None,
    department_id: int | None,
) -> list[tuple[AttendanceEvent, Employee, Department | None]]:
    if end_date < start_date:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="end_date must be >= start_date",
        )

    start_dt, end_dt = _date_range_bounds_utc(start_date, end_date)

    stmt = (
        select(AttendanceEvent, Employee, Department)
        .join(Employee, AttendanceEvent.employee_id == Employee.id)
        .outerjoin(Department, Employee.department_id == Department.id)
        .where(
            AttendanceEvent.ts_utc >= start_dt,
            AttendanceEvent.ts_utc < end_dt,
            AttendanceEvent.deleted_at.is_(None),
        )
        .order_by(AttendanceEvent.ts_utc.asc(), AttendanceEvent.id.asc())
    )
    if employee_id is not None:
        stmt = stmt.where(AttendanceEvent.employee_id == employee_id)
    if department_id is not None:
        stmt = stmt.where(Employee.department_id == department_id)

    return list(db.execute(stmt).all())


def _true_flag_names(flags: dict[str, object] | None) -> str:
    if not flags:
        return "-"
    names = [key for key, value in flags.items() if isinstance(value, bool) and value]
    if not names:
        return "-"
    return ", ".join(sorted(names))


def _write_range_sheet_rows(
    ws: Worksheet,
    *,
    rows: list[tuple[AttendanceEvent, Employee, Department | None]],
) -> None:
    _merge_title(ws, 1, "PUANTAJ TARIH ARALIGI RAPORU", end_col=11)
    ws.append(["Kayıt Sayısı", len(rows)])
    ws.append(["Rapor Üretim (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")])
    _append_spacer_row(ws)
    _style_metadata_rows(ws, start_row=2, end_row=3)

    header_row = ws.max_row + 1
    ws.append(
        [
            "Tarih",
            "Saat (UTC)",
            "Çalışan ID",
            "Çalışan Adı",
            "Departman",
            "Tip",
            "Konum Durumu",
            "Enlem",
            "Boylam",
            "Doğruluk (m)",
            "Bayraklar",
        ]
    )
    _style_header(ws, header_row)

    for event, employee, department in rows:
        ts_value = _to_excel_datetime(event.ts_utc)
        event_date = ts_value.date() if ts_value else None
        ws.append(
            [
                event_date,
                ts_value,
                event.employee_id,
                employee.full_name,
                department.name if department is not None else "-",
                event.type.value,
                event.location_status.value,
                event.lat,
                event.lon,
                event.accuracy_m,
                _true_flag_names(event.flags or {}),
            ]
        )

    data_start = header_row + 1
    data_end = ws.max_row
    _style_table_region(
        ws,
        header_row=header_row,
        data_start_row=data_start,
        data_end_row=data_end,
        status_col_name="Konum Durumu",
        flags_col_name="Bayraklar",
    )

    if ws.max_row >= data_start:
        for row in ws.iter_rows(min_row=data_start, max_row=data_end):
            if row[0].value is not None:
                row[0].number_format = "yyyy-mm-dd"
            if row[1].value is not None:
                row[1].number_format = "hh:mm"
    _auto_width(ws)
    _finalize_visual_scope(ws, used_max_col=11)
    _apply_print_layout(ws, header_row=header_row)


def build_puantaj_range_xlsx_bytes(
    db: Session,
    *,
    start_date: date,
    end_date: date,
    mode: RangeSheetMode,
    department_id: int | None = None,
    employee_id: int | None = None,
) -> bytes:
    if mode not in {"consolidated", "employee_sheets", "department_sheets"}:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Invalid range export mode",
        )

    wb = Workbook()
    _build_date_range_export(
        db,
        wb,
        start_date=start_date,
        end_date=end_date,
        employee_id=employee_id,
        department_id=department_id,
    )

    if mode == "consolidated":
        _build_readme_sheet(wb, report_title="Tarih Araligi Raporu")
        _apply_workbook_branding(wb)
        stream = BytesIO()
        wb.save(stream)
        return stream.getvalue()

    rows = _fetch_date_range_rows(
        db,
        start_date=start_date,
        end_date=end_date,
        employee_id=employee_id,
        department_id=department_id,
    )

    if mode == "employee_sheets":
        grouped: dict[tuple[int, str], list[tuple[AttendanceEvent, Employee, Department | None]]] = defaultdict(list)
        for row in rows:
            event, employee, _department = row
            grouped[(event.employee_id, employee.full_name)].append(row)

        dashboard_links: list[tuple[int, str, str]] = []
        for (emp_id, full_name), group_rows in sorted(grouped.items(), key=lambda item: item[0][0]):
            ws = wb.create_sheet(_safe_sheet_title(f"EMP_{emp_id}_{full_name}", f"Emp{emp_id}"))
            _write_range_sheet_rows(ws, rows=group_rows)
            dashboard_links.append((emp_id, full_name, ws.title))

        if dashboard_links and "DASHBOARD" in wb.sheetnames:
            ws_dashboard = wb["DASHBOARD"]
            ws_dashboard.append([])
            nav_header = ws_dashboard.max_row + 1
            ws_dashboard.append(["Çalışan", "Sayfa"])
            _style_header(ws_dashboard, nav_header)
            nav_data_start = ws_dashboard.max_row + 1
            for emp_id, full_name, sheet_title in dashboard_links:
                ws_dashboard.append([f"#{emp_id} - {full_name}", sheet_title])
                _set_internal_sheet_link(ws_dashboard.cell(row=ws_dashboard.max_row, column=1), sheet_title)
                _set_internal_sheet_link(ws_dashboard.cell(row=ws_dashboard.max_row, column=2), sheet_title)
            nav_data_end = ws_dashboard.max_row
            _style_table_region(
                ws_dashboard,
                header_row=nav_header,
                data_start_row=nav_data_start,
                data_end_row=nav_data_end,
                status_col_name="",
                flags_col_name="",
            )
            _auto_width(ws_dashboard)
            _apply_print_layout(ws_dashboard)
    elif mode == "department_sheets":
        grouped: dict[str, list[tuple[AttendanceEvent, Employee, Department | None]]] = defaultdict(list)
        for row in rows:
            _event, _employee, department = row
            key = department.name if department is not None else "Atanmamış"
            grouped[key].append(row)

        for dep_name, group_rows in sorted(grouped.items(), key=lambda item: item[0].lower()):
            ws = wb.create_sheet(_safe_sheet_title(f"DEP_{dep_name}", "Departman"))
            _write_range_sheet_rows(ws, rows=group_rows)

    _build_readme_sheet(wb, report_title="Tarih Araligi Raporu")
    _apply_workbook_branding(wb)
    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
