from __future__ import annotations

import unittest
from collections.abc import Generator
from datetime import date, datetime, timezone
from io import BytesIO
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.db import get_db
from app.main import app
from app.models import (
    AttendanceEvent,
    AttendanceType,
    Department,
    DepartmentShift,
    Device,
    DeviceInvite,
    Employee,
    LocationStatus,
    ManualDayOverride,
    Region,
)
from app.schemas import MonthlyEmployeeDay, MonthlyEmployeeResponse, MonthlyEmployeeTotals
from app.security import require_admin
from app.services.monthly import calculate_employee_monthly


class _ScalarRows:
    def __init__(self, rows):
        self._rows = rows

    def all(self):
        return self._rows


def _override_get_db(fake_db):
    def _override() -> Generator[object, None, None]:
        yield fake_db

    return _override


def _super_admin_claims() -> dict[str, object]:
    return {
        "sub": "admin",
        "username": "admin",
        "role": "admin",
        "iat": 0,
        "exp": 9999999999,
        "jti": "test-admin",
        "is_super_admin": True,
        "permissions": {},
    }


class _FakeExportDB:
    def __init__(self, employee: Employee):
        self._employee = employee
        self.audit_rows: list[object] = []

    def get(self, model, pk):  # type: ignore[no-untyped-def]
        if model is Employee and pk == self._employee.id:
            return self._employee
        return None

    def add(self, obj: object) -> None:
        self.audit_rows.append(obj)

    def commit(self) -> None:
        return

    def rollback(self) -> None:
        return


class _FakeAllExportDB:
    def __init__(self, employees: list[Employee]):
        self._employees = employees
        self.audit_rows: list[object] = []

    def scalars(self, _statement):  # type: ignore[no-untyped-def]
        return _ScalarRows(self._employees)

    def scalar(self, _statement):  # type: ignore[no-untyped-def]
        return None

    def add(self, obj: object) -> None:
        self.audit_rows.append(obj)

    def commit(self) -> None:
        return

    def rollback(self) -> None:
        return


class _FakeMonthlyDB:
    def __init__(self, scalar_values, scalars_values):
        self._scalar_values = list(scalar_values)
        self._scalars_values = list(scalars_values)

    def scalar(self, _statement):  # type: ignore[no-untyped-def]
        if not self._scalar_values:
            return None
        return self._scalar_values.pop(0)

    def scalars(self, _statement):  # type: ignore[no-untyped-def]
        if not self._scalars_values:
            return _ScalarRows([])
        return _ScalarRows(self._scalars_values.pop(0))


class _FakeEmployeeListDB:
    def __init__(self, active_rows, all_rows):
        self.active_rows = active_rows
        self.all_rows = all_rows

    def scalars(self, statement):  # type: ignore[no-untyped-def]
        stmt_text = str(statement)
        if "employees.is_active IS true" in stmt_text:
            return _ScalarRows(self.active_rows)
        return _ScalarRows(self.all_rows)


class _FakeEmployeeLifecycleDB:
    def __init__(self, employees):
        self.employees = {employee.id: employee for employee in employees}
        self.audit_rows: list[object] = []

    def scalars(self, statement):  # type: ignore[no-untyped-def]
        stmt_text = str(statement)
        rows = list(self.employees.values())
        if "employees.is_active IS true" in stmt_text:
            rows = [employee for employee in rows if employee.is_active]
        elif "employees.is_active IS false" in stmt_text:
            rows = [employee for employee in rows if not employee.is_active]
        return _ScalarRows(rows)

    def get(self, model, pk):  # type: ignore[no-untyped-def]
        if model is Employee:
            return self.employees.get(pk)
        return None

    def add(self, obj: object) -> None:
        self.audit_rows.append(obj)

    def commit(self) -> None:
        return

    def refresh(self, _obj: object) -> None:
        return

    def rollback(self) -> None:
        return


class _FakeEmployeeDeviceOverviewDB:
    def __init__(self, employees):
        self.employees = employees

    def scalars(self, statement):  # type: ignore[no-untyped-def]
        stmt_text = str(statement)
        rows = list(self.employees)
        if "employees.is_active IS true" in stmt_text:
            rows = [employee for employee in rows if employee.is_active]
        return _ScalarRows(rows)


class _FakeEmployeeDepartmentDB:
    def __init__(self, employee, departments, shifts):
        self.employee = employee
        self.departments = {department.id: department for department in departments}
        self.shifts = {shift.id: shift for shift in shifts}
        self.audit_rows: list[object] = []

    def get(self, model, pk):  # type: ignore[no-untyped-def]
        if model is Employee and pk == self.employee.id:
            return self.employee
        if model is Department:
            return self.departments.get(pk)
        if model is DepartmentShift:
            return self.shifts.get(pk)
        return None

    def add(self, obj: object) -> None:
        self.audit_rows.append(obj)

    def commit(self) -> None:
        return

    def refresh(self, _obj: object) -> None:
        return


class _FakeDepartmentManageDB:
    def __init__(self, department):
        self.department = department
        self.audit_rows: list[object] = []

    def get(self, model, pk):  # type: ignore[no-untyped-def]
        if model is Department and pk == self.department.id:
            return self.department
        return None

    def add(self, obj: object) -> None:
        self.audit_rows.append(obj)

    def scalar(self, _statement):  # type: ignore[no-untyped-def]
        return self.department

    def commit(self) -> None:
        return

    def refresh(self, _obj: object) -> None:
        return

    def rollback(self) -> None:
        return


class _FakeDepartmentRegionListDB:
    def __init__(self, departments):
        self.departments = departments

    def scalars(self, statement):  # type: ignore[no-untyped-def]
        stmt_text = str(statement)
        rows = list(self.departments)
        if "departments.region_id =" in stmt_text:
            rows = [item for item in rows if item.region_id == 1]
        return _ScalarRows(rows)


class AdminPuantajFeatureTests(unittest.TestCase):
    def tearDown(self) -> None:
        app.dependency_overrides.clear()

    def test_xlsx_export_endpoint_returns_valid_non_empty_file(self) -> None:
        employee = Employee(id=1, full_name="Test User", department_id=None, is_active=True)
        fake_db = _FakeExportDB(employee)
        app.dependency_overrides[get_db] = _override_get_db(fake_db)
        app.dependency_overrides[require_admin] = lambda: _super_admin_claims()

        fake_report = MonthlyEmployeeResponse(
            employee_id=1,
            year=2026,
            month=2,
            days=[
                MonthlyEmployeeDay(
                    date=date(2026, 2, 1),
                    status="OK",
                    check_in=datetime(2026, 2, 1, 9, 0, tzinfo=timezone.utc),
                    check_out=datetime(2026, 2, 1, 18, 0, tzinfo=timezone.utc),
                    worked_minutes=480,
                    overtime_minutes=0,
                    flags=[],
                )
            ],
            totals=MonthlyEmployeeTotals(
                worked_minutes=480,
                overtime_minutes=0,
                incomplete_days=0,
            ),
            worked_minutes_net=480,
            weekly_totals=[],
            annual_overtime_used_minutes=0,
            annual_overtime_remaining_minutes=16200,
            annual_overtime_cap_exceeded=False,
            labor_profile=None,
        )

        client = TestClient(app)
        with patch("app.services.exports.calculate_employee_monthly", return_value=fake_report):
            response = client.get(
                "/api/admin/exports/puantaj.xlsx?mode=employee&employee_id=1&year=2026&month=2"
            )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.headers.get("content-type", ""),
        )

        wb = load_workbook(BytesIO(response.content))
        self.assertGreater(len(wb.sheetnames), 0)
        first_sheet = wb[wb.sheetnames[0]]
        self.assertGreater(first_sheet.max_row, 1)
        self.assertGreater(first_sheet.max_column, 1)

    def test_xlsx_export_formats_dashboard_and_duration_columns(self) -> None:
        employee = Employee(id=1, full_name="Test User", department_id=None, is_active=True)
        fake_db = _FakeExportDB(employee)
        app.dependency_overrides[get_db] = _override_get_db(fake_db)
        app.dependency_overrides[require_admin] = lambda: _super_admin_claims()

        fake_report = MonthlyEmployeeResponse(
            employee_id=1,
            year=2026,
            month=2,
            days=[
                MonthlyEmployeeDay(
                    date=date(2026, 2, 12),
                    status="OK",
                    check_in=datetime(2026, 2, 12, 10, 47, tzinfo=timezone.utc),
                    check_out=datetime(2026, 2, 12, 16, 42, tzinfo=timezone.utc),
                    worked_minutes=355,
                    overtime_minutes=175,
                    flags=[],
                )
            ],
            totals=MonthlyEmployeeTotals(
                worked_minutes=355,
                overtime_minutes=175,
                incomplete_days=0,
            ),
            worked_minutes_net=355,
            weekly_totals=[],
            annual_overtime_used_minutes=0,
            annual_overtime_remaining_minutes=16200,
            annual_overtime_cap_exceeded=False,
            labor_profile=None,
        )

        client = TestClient(app)
        with patch("app.services.exports.calculate_employee_monthly", return_value=fake_report):
            response = client.get(
                "/api/admin/exports/puantaj.xlsx?mode=employee&employee_id=1&year=2026&month=2"
            )

        self.assertEqual(response.status_code, 200)
        wb = load_workbook(BytesIO(response.content))

        dashboard = wb["DASHBOARD"]
        self.assertEqual(dashboard["B5"].number_format, "0")
        self.assertEqual(dashboard["B10"].number_format, "0")
        for cell_ref in ("B6", "B7", "B8", "B9"):
            self.assertEqual(dashboard[cell_ref].number_format, "[h]:mm")

        overtime_header_row = None
        for row_idx in range(1, dashboard.max_row + 1):
            if dashboard.cell(row=row_idx, column=4).value == "Yasal Fazla Mesai [h:mm]":
                overtime_header_row = row_idx
                break
        self.assertIsNotNone(overtime_header_row)
        self.assertEqual(dashboard.cell(row=overtime_header_row + 1, column=4).number_format, "[h]:mm")

        column_a_values = [dashboard.cell(row=row_idx, column=1).value for row_idx in range(1, dashboard.max_row + 1)]
        self.assertNotIn("Çalışan/Sayfa", column_a_values)

        daily_sheet = wb["DAILY_1"]
        day_row = None
        target_date = date(2026, 2, 12)
        for row_idx in range(1, daily_sheet.max_row + 1):
            value = daily_sheet.cell(row=row_idx, column=1).value
            if isinstance(value, datetime) and value.date() == target_date:
                day_row = row_idx
                break
            if isinstance(value, date) and value == target_date:
                day_row = row_idx
                break
        self.assertIsNotNone(day_row)

        for col_idx in (5, 6, 7, 9, 11, 12):
            self.assertEqual(daily_sheet.cell(row=day_row, column=col_idx).number_format, "[h]:mm")
        for col_idx in (8, 10, 13):
            self.assertNotEqual(daily_sheet.cell(row=day_row, column=col_idx).number_format, "[h]:mm")

    def test_xlsx_export_all_mode_uses_clean_turkish_labels(self) -> None:
        department = Department(id=1, name="ARGE")
        employee_1 = Employee(id=1, full_name="Hüseyincan Orman", department_id=1, is_active=True)
        employee_2 = Employee(id=2, full_name="Birtan Başkaya", department_id=1, is_active=True)
        employee_1.department = department
        employee_2.department = department

        fake_db = _FakeAllExportDB([employee_1, employee_2])
        app.dependency_overrides[get_db] = _override_get_db(fake_db)
        app.dependency_overrides[require_admin] = lambda: _super_admin_claims()

        fake_report = MonthlyEmployeeResponse(
            employee_id=1,
            year=2026,
            month=2,
            days=[
                MonthlyEmployeeDay(
                    date=date(2026, 2, 12),
                    status="OK",
                    check_in=datetime(2026, 2, 12, 10, 47, tzinfo=timezone.utc),
                    check_out=datetime(2026, 2, 12, 16, 42, tzinfo=timezone.utc),
                    worked_minutes=355,
                    overtime_minutes=175,
                    flags=[],
                )
            ],
            totals=MonthlyEmployeeTotals(
                worked_minutes=355,
                overtime_minutes=175,
                incomplete_days=0,
            ),
            worked_minutes_net=355,
            weekly_totals=[],
            annual_overtime_used_minutes=0,
            annual_overtime_remaining_minutes=16200,
            annual_overtime_cap_exceeded=False,
            labor_profile=None,
        )

        client = TestClient(app)
        with patch("app.services.exports.calculate_employee_monthly", return_value=fake_report):
            response = client.get("/api/admin/exports/puantaj.xlsx?mode=all&year=2026&month=2")

        self.assertEqual(response.status_code, 200)
        wb = load_workbook(BytesIO(response.content))
        dashboard = wb["DASHBOARD"]

        self.assertEqual(dashboard["A1"].value, "Tüm Çalışanlar Özeti - YÖNETİM DASHBOARD")
        self.assertEqual(dashboard["A2"].value, "Rapor Üretim (UTC)")
        self.assertEqual(dashboard["A12"].value, "Sıra")
        self.assertEqual(dashboard["B12"].value, "Çalışan Navigasyon")

        daily_sheet = wb["Hüseyincan Orman (1)"]
        self.assertEqual(daily_sheet["E7"].value, "Çalışma Süresi")
        self.assertEqual(daily_sheet["N7"].value, "Gün Tipi")
        self.assertEqual(daily_sheet["O7"].value, "Çalışıldı mı")

        mojibake_markers = ("Ã", "Â", "Ä", "Å")
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 40), values_only=True):
                for value in row:
                    if not isinstance(value, str):
                        continue
                    self.assertFalse(any(marker in value for marker in mojibake_markers))

    def test_new_employee_monthly_export_endpoint_returns_valid_xlsx(self) -> None:
        employee = Employee(id=1, full_name="Test User", department_id=None, is_active=True)
        fake_db = _FakeExportDB(employee)
        app.dependency_overrides[get_db] = _override_get_db(fake_db)
        app.dependency_overrides[require_admin] = lambda: _super_admin_claims()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Puantaj"
        sheet.append(["Tarih", "Net Süre"])
        sheet.append(["2026-02-01", "08:00"])
        stream = BytesIO()
        workbook.save(stream)

        client = TestClient(app)
        with patch("app.routers.admin.build_puantaj_xlsx_bytes", return_value=stream.getvalue()):
            response = client.get("/api/admin/export/employee-monthly?employee_id=1&year=2026&month=2")

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.headers.get("content-type", ""),
        )
        wb = load_workbook(BytesIO(response.content))
        self.assertGreater(len(wb.sheetnames), 0)
        self.assertGreater(wb[wb.sheetnames[0]].max_row, 1)

    def test_new_date_range_export_endpoint_returns_valid_xlsx(self) -> None:
        employee = Employee(id=1, full_name="Test User", department_id=None, is_active=True)
        fake_db = _FakeExportDB(employee)
        app.dependency_overrides[get_db] = _override_get_db(fake_db)
        app.dependency_overrides[require_admin] = lambda: _super_admin_claims()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Gunluk Ozet"
        sheet.append(["Tarih", "Durum"])
        sheet.append(["2026-02-01", "OK"])
        stream = BytesIO()
        workbook.save(stream)

        client = TestClient(app)
        with patch("app.routers.admin.build_puantaj_xlsx_bytes", return_value=stream.getvalue()):
            response = client.get("/api/admin/export/date-range?start=2026-02-01&end=2026-02-28")

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.headers.get("content-type", ""),
        )
        wb = load_workbook(BytesIO(response.content))
        self.assertGreater(len(wb.sheetnames), 0)
        self.assertGreater(wb[wb.sheetnames[0]].max_row, 1)

    def test_new_puantaj_range_export_endpoint_returns_valid_xlsx(self) -> None:
        employee = Employee(id=1, full_name="Test User", department_id=None, is_active=True)
        fake_db = _FakeExportDB(employee)
        app.dependency_overrides[get_db] = _override_get_db(fake_db)
        app.dependency_overrides[require_admin] = lambda: _super_admin_claims()

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Konsolide"
        sheet.append(["Tarih", "Tip"])
        sheet.append(["2026-02-01", "IN"])
        stream = BytesIO()
        workbook.save(stream)

        client = TestClient(app)
        with patch("app.routers.admin.build_puantaj_range_xlsx_bytes", return_value=stream.getvalue()):
            response = client.get(
                "/api/admin/export/puantaj-range.xlsx?start_date=2026-02-01&end_date=2026-02-28&mode=consolidated"
            )

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.headers.get("content-type", ""),
        )
        wb = load_workbook(BytesIO(response.content))
        self.assertGreater(len(wb.sheetnames), 0)
        self.assertGreater(wb[wb.sheetnames[0]].max_row, 1)

    def test_manual_override_changes_monthly_day_output(self) -> None:
        employee = Employee(id=7, full_name="Override User", department_id=None, is_active=True)
        event_in = AttendanceEvent(
            employee_id=7,
            device_id=1,
            type=AttendanceType.IN,
            ts_utc=datetime(2026, 2, 1, 8, 0, tzinfo=timezone.utc),
            location_status=LocationStatus.NO_LOCATION,
            flags={},
        )
        event_out = AttendanceEvent(
            employee_id=7,
            device_id=1,
            type=AttendanceType.OUT,
            ts_utc=datetime(2026, 2, 1, 16, 0, tzinfo=timezone.utc),
            location_status=LocationStatus.NO_LOCATION,
            flags={},
        )
        override = ManualDayOverride(
            employee_id=7,
            day_date=date(2026, 2, 1),
            in_ts=datetime(2026, 2, 1, 9, 30, tzinfo=timezone.utc),
            out_ts=datetime(2026, 2, 1, 18, 30, tzinfo=timezone.utc),
            is_absent=False,
            note="IK duzeltmesi",
            created_by="admin",
        )
        fake_db = _FakeMonthlyDB(
            scalar_values=[employee, None, None],
            scalars_values=[
                [event_in, event_out],
                [],
                [override],
            ],
        )

        report = calculate_employee_monthly(fake_db, employee_id=7, year=2026, month=2)
        first_day = report.days[0]

        self.assertEqual(first_day.status, "OK")
        self.assertEqual(first_day.check_in, override.in_ts)
        self.assertEqual(first_day.check_out, override.out_ts)
        self.assertEqual(first_day.worked_minutes, 480)
        self.assertIn("MANUAL_OVERRIDE", first_day.flags)

    def test_default_employee_list_excludes_inactive(self) -> None:
        active_employee = Employee(id=1, full_name="Aktif Kisi", department_id=None, is_active=True)
        inactive_employee = Employee(id=2, full_name="Pasif Kisi", department_id=None, is_active=False)

        fake_db = _FakeEmployeeListDB(
            active_rows=[active_employee],
            all_rows=[active_employee, inactive_employee],
        )
        app.dependency_overrides[get_db] = _override_get_db(fake_db)
        app.dependency_overrides[require_admin] = lambda: _super_admin_claims()
        client = TestClient(app)

        active_response = client.get("/admin/employees")
        all_response = client.get("/admin/employees?include_inactive=true")

        self.assertEqual(active_response.status_code, 200)
        self.assertEqual(all_response.status_code, 200)

        active_items = active_response.json()
        all_items = all_response.json()

        self.assertEqual(len(active_items), 1)
        self.assertEqual(active_items[0]["id"], 1)
        self.assertEqual(len(all_items), 2)

    def test_deactivate_then_reactivate_employee_with_include_inactive(self) -> None:
        employee = Employee(id=9, full_name="Yeniden Aktif", department_id=None, is_active=True)
        fake_db = _FakeEmployeeLifecycleDB([employee])
        app.dependency_overrides[get_db] = _override_get_db(fake_db)
        app.dependency_overrides[require_admin] = lambda: _super_admin_claims()
        client = TestClient(app)

        deactivate_response = client.patch(
            "/api/admin/employees/9/active",
            json={"is_active": False},
        )
        self.assertEqual(deactivate_response.status_code, 200)
        self.assertFalse(deactivate_response.json()["is_active"])

        default_list_response = client.get("/admin/employees")
        self.assertEqual(default_list_response.status_code, 200)
        self.assertEqual(default_list_response.json(), [])

        include_inactive_response = client.get("/admin/employees?include_inactive=true")
        self.assertEqual(include_inactive_response.status_code, 200)
        self.assertEqual(len(include_inactive_response.json()), 1)
        self.assertFalse(include_inactive_response.json()[0]["is_active"])

        reactivate_response = client.patch(
            "/api/admin/employees/9/active",
            json={"is_active": True},
        )
        self.assertEqual(reactivate_response.status_code, 200)
        self.assertTrue(reactivate_response.json()["is_active"])

        final_default_list = client.get("/admin/employees")
        self.assertEqual(final_default_list.status_code, 200)
        self.assertEqual(len(final_default_list.json()), 1)
        self.assertTrue(final_default_list.json()[0]["is_active"])

    def test_employee_device_overview_returns_token_and_device_counts(self) -> None:
        employee = Employee(id=3, full_name="Cihaz Test", department_id=1, is_active=True)
        employee.department = Department(id=1, name="Operasyon")
        employee.devices = [
            Device(
                id=11,
                employee_id=3,
                device_fingerprint="device-a",
                is_active=True,
                created_at=datetime(2026, 2, 10, 10, 0, tzinfo=timezone.utc),
            ),
            Device(
                id=12,
                employee_id=3,
                device_fingerprint="device-b",
                is_active=False,
                created_at=datetime(2026, 2, 9, 10, 0, tzinfo=timezone.utc),
            ),
        ]
        employee.device_invites = [
            DeviceInvite(
                id=21,
                employee_id=3,
                token="used-token",
                expires_at=datetime(2026, 2, 28, 0, 0, tzinfo=timezone.utc),
                is_used=True,
                created_at=datetime(2026, 2, 1, 0, 0, tzinfo=timezone.utc),
            ),
            DeviceInvite(
                id=22,
                employee_id=3,
                token="pending-token",
                expires_at=datetime(2099, 2, 28, 0, 0, tzinfo=timezone.utc),
                is_used=False,
                created_at=datetime(2026, 2, 2, 0, 0, tzinfo=timezone.utc),
            ),
            DeviceInvite(
                id=23,
                employee_id=3,
                token="expired-token",
                expires_at=datetime(2020, 2, 28, 0, 0, tzinfo=timezone.utc),
                is_used=False,
                created_at=datetime(2026, 2, 3, 0, 0, tzinfo=timezone.utc),
            ),
        ]

        fake_db = _FakeEmployeeDeviceOverviewDB([employee])
        app.dependency_overrides[get_db] = _override_get_db(fake_db)
        app.dependency_overrides[require_admin] = lambda: _super_admin_claims()
        client = TestClient(app)

        response = client.get("/api/admin/employee-device-overview")
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(len(payload), 1)
        self.assertEqual(payload[0]["employee_id"], 3)
        self.assertEqual(payload[0]["token_total"], 3)
        self.assertEqual(payload[0]["token_used"], 1)
        self.assertEqual(payload[0]["token_pending"], 1)
        self.assertEqual(payload[0]["token_expired"], 1)
        self.assertEqual(len(payload[0]["devices"]), 2)

    @patch("app.routers.admin.log_audit")
    @patch("app.routers.admin.get_admin_recovery_snapshot")
    def test_employee_device_overview_can_include_recovery_secrets(
        self,
        mock_get_admin_recovery_snapshot,
        _mock_log_audit,
    ) -> None:
        employee = Employee(id=4, full_name="Recovery Test", department_id=1, is_active=True)
        employee.department = Department(id=1, name="Operasyon")
        employee.devices = [
            Device(
                id=31,
                employee_id=4,
                device_fingerprint="device-recovery",
                is_active=True,
                created_at=datetime(2026, 2, 11, 9, 30, tzinfo=timezone.utc),
            )
        ]
        employee.device_invites = []
        mock_get_admin_recovery_snapshot.return_value = {
            "recovery_ready": True,
            "recovery_code_active_count": 7,
            "recovery_expires_at": datetime(2026, 12, 31, 18, 0, tzinfo=timezone.utc),
            "recovery_pin_updated_at": datetime(2026, 2, 11, 9, 31, tzinfo=timezone.utc),
            "recovery_pin_plain": "123456",
            "recovery_code_entries": [{"code": "AB3D-9K2M", "status": "ACTIVE"}],
        }

        fake_db = _FakeEmployeeDeviceOverviewDB([employee])
        app.dependency_overrides[get_db] = _override_get_db(fake_db)
        app.dependency_overrides[require_admin] = lambda: _super_admin_claims()
        client = TestClient(app)

        response = client.get("/api/admin/employee-device-overview?employee_id=4&include_recovery_secrets=true")
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(len(payload), 1)
        self.assertEqual(len(payload[0]["devices"]), 1)
        device = payload[0]["devices"][0]
        self.assertTrue(device["recovery_ready"])
        self.assertEqual(device["recovery_code_active_count"], 7)
        self.assertEqual(device["recovery_pin_plain"], "123456")
        self.assertEqual(device["recovery_code_entries"][0]["status"], "ACTIVE")

    def test_update_employee_department_clears_shift_if_department_changes(self) -> None:
        employee = Employee(id=6, full_name="Departman Test", region_id=1, department_id=1, shift_id=7, is_active=True)
        old_department = Department(id=1, name="Eski Departman", region_id=1)
        new_department = Department(id=2, name="Yeni Departman", region_id=2)
        old_shift = DepartmentShift(
            id=7,
            department_id=1,
            name="Sabah",
            start_time_local=datetime(2026, 1, 1, 9, 0).time(),
            end_time_local=datetime(2026, 1, 1, 17, 0).time(),
            break_minutes=60,
            is_active=True,
        )

        fake_db = _FakeEmployeeDepartmentDB(employee, [old_department, new_department], [old_shift])
        app.dependency_overrides[get_db] = _override_get_db(fake_db)
        app.dependency_overrides[require_admin] = lambda: _super_admin_claims()
        client = TestClient(app)

        response = client.patch("/api/admin/employees/6/department", json={"department_id": 2})
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["department_id"], 2)
        self.assertEqual(payload["region_id"], 2)
        self.assertIsNone(payload["shift_id"])

    def test_update_department_endpoint_updates_department_name(self) -> None:
        department = Department(id=5, name="Eski Departman")
        fake_db = _FakeDepartmentManageDB(department)
        app.dependency_overrides[get_db] = _override_get_db(fake_db)
        app.dependency_overrides[require_admin] = lambda: _super_admin_claims()
        client = TestClient(app)

        response = client.patch("/admin/departments/5", json={"name": "Yeni Departman"})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["id"], 5)
        self.assertEqual(payload["name"], "Yeni Departman")

    def test_list_departments_can_filter_by_region(self) -> None:
        region_1 = Region(id=1, name="Istanbul", is_active=True)
        region_2 = Region(id=2, name="Ankara", is_active=True)
        department_1 = Department(id=11, name="Satis", region_id=1)
        department_1.region = region_1
        department_2 = Department(id=12, name="Operasyon", region_id=2)
        department_2.region = region_2

        fake_db = _FakeDepartmentRegionListDB([department_1, department_2])
        app.dependency_overrides[get_db] = _override_get_db(fake_db)
        app.dependency_overrides[require_admin] = lambda: _super_admin_claims()
        client = TestClient(app)

        response = client.get("/admin/departments?region_id=1")
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(len(payload), 1)
        self.assertEqual(payload[0]["id"], 11)
        self.assertEqual(payload[0]["region_id"], 1)


if __name__ == "__main__":
    unittest.main()
